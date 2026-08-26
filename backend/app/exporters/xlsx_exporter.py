import os
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class XlsxExporter:
    """
    Exportador extendido de documentos a hojas de cálculo Microsoft Excel (.xlsx).
    Organiza la información estructurada y campos en formato matricial.
    """

    def __init__(self):
        pass

    def export_to_file(
        self,
        document_title: str,
        fields: List[Dict[str, Any]],
        output_path: str,
        images_to_embed: Optional[List[str]] = None
    ) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulario Digitalizado"

        # Estilos
        font_title = Font(name="Calibri", size=14, bold=True, color="1E293B")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11, color="334155")
        
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Título
        ws.merge_cells("A1:D1")
        cell_title = ws["A1"]
        cell_title.value = document_title or "Formulario Digitalizado - DocScan"
        cell_title.font = font_title
        cell_title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Encabezados de tabla
        headers = ["ID Campo", "Etiqueta / Campo", "Tipo de Dato", "Valor / Espacio Editable"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[3].height = 25

        # Filas de datos
        row_idx = 4
        for idx, field in enumerate(fields, 1):
            etiqueta = field.get("etiqueta", f"Campo {idx}")
            tipo = field.get("tipo_campo", "texto").capitalize()
            
            cell_id = ws.cell(row=row_idx, column=1, value=idx)
            cell_lbl = ws.cell(row=row_idx, column=2, value=etiqueta)
            cell_tipo = ws.cell(row=row_idx, column=3, value=tipo)
            cell_val = ws.cell(row=row_idx, column=4, value="") # Espacio vacío para diligenciar

            for c in [cell_id, cell_lbl, cell_tipo, cell_val]:
                c.font = font_data
                c.border = thin_border
                if row_idx % 2 == 0:
                    c.fill = fill_zebra

            ws.row_dimensions[row_idx].height = 22
            row_idx += 1

        # Ancho de columnas automático
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40

        # Inserción de Imágenes si se proveen
        if images_to_embed:
            from openpyxl.drawing.image import Image as OpenPyXLImage
            ws.cell(row=row_idx + 2, column=1, value="Imágenes Incrustadas:").font = Font(bold=True)
            for idx_img, img_path in enumerate(images_to_embed):
                if os.path.exists(img_path):
                    try:
                        img = OpenPyXLImage(img_path)
                        img.width = 180
                        img.height = 100
                        ws.add_image(img, f"B{row_idx + 4 + (idx_img * 6)}")
                    except Exception as e:
                        pass

        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path
