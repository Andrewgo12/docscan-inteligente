import openpyxl
from typing import Dict, Any, List

class TableCellExtractor:
    """Extractor especializado de celdas, filas y tablas de hojas de cálculo."""

    def extract_cells_and_headers(self, excel_path: str) -> List[Dict[str, Any]]:
        sheets_data = []
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                sheets_data.append({
                    "sheet_name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "rows_sample": [list(r) for r in rows[:10] if r]
                })
        except Exception as e:
            sheets_data.append({"sheet_name": "Sheet1", "error": str(e)})
        return sheets_data
