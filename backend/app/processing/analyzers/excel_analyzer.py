import os
import zipfile
from typing import Dict, Any, List
import openpyxl

EXTRACTED_IMG_DIR = "backend/storage/uploads/extracted_images"
os.makedirs(EXTRACTED_IMG_DIR, exist_ok=True)

class ExcelAnalyzer:
    """
    Analizador Especializado de Hojas de Cálculo Microsoft Excel (.xlsx).
    """

    def extract_embedded_images(self, file_path: str) -> List[str]:
        extracted_paths = []
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                for member in z.namelist():
                    if member.startswith('xl/media/') and not member.endswith('/'):
                        img_filename = os.path.basename(member)
                        out_path = os.path.join(EXTRACTED_IMG_DIR, f"{os.path.basename(file_path)}_{img_filename}")
                        with open(out_path, 'wb') as f_out:
                            f_out.write(z.read(member))
                        extracted_paths.append(out_path)
        except Exception:
            pass
        return extracted_paths

    def analyze(self, file_path: str) -> Dict[str, Any]:
        size_bytes = os.path.getsize(file_path)
        extracted_images = self.extract_embedded_images(file_path)
        sheets_detail = []
        total_cells = 0

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                non_empty_cells = sum(1 for r in rows[:1000] for c in r if c is not None)
                total_cells += non_empty_cells

                sheets_detail.append({
                    "sheet_name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "rows_count": ws.max_row,
                    "columns_count": ws.max_column,
                    "non_empty_cells": non_empty_cells
                })
        except Exception as e:
            sheets_detail.append({"sheet_name": "Sheet1", "rows_count": 0, "columns_count": 0, "error": str(e)})

        executive_summary = (
            f"Excel analizado con {len(sheets_detail)} hoja(s). Total celdas: {total_cells}. "
            f"Imágenes incrustadas: {len(extracted_images)}."
        )

        return {
            "format": "excel",
            "file_type": "office_excel",
            "file_name": os.path.basename(file_path),
            "size_bytes": size_bytes,
            "total_sheets": len(sheets_detail),
            "sheets_count": len(sheets_detail),
            "sheets": sheets_detail,
            "sheets_detail": sheets_detail,
            "total_cells": total_cells,
            "total_extracted_images": len(extracted_images),
            "extracted_images_paths": extracted_images,
            "executive_summary": executive_summary,
            "summary": executive_summary
        }
